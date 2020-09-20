# -*- coding: utf-8 -*-
from itertools import izip

from openpyxl import load_workbook

from constants import REQUIRED_COLUMNS, INPUT_SPREADSHEET_MAX_ROWS


def sanitize_column_name(column_name):
    """Convierte algunas variantes de nombres de columnas a su forma 'canónica'"""
    if column_name.lower() == u'teléfono':
        return 'telefono'
    else:
        return column_name


def validate_required_columns(columns):
    """Se valida que el archivo incluya las columnas obligatorias"""
    for column_name in REQUIRED_COLUMNS:
        if column_name not in columns:
            raise Exception('El archivo de entrada no contiene la columna "{}"'
                            .format(column_name))


def generate_columns_dict(filename):
    """Recibe la ruta a un .xlsx y devuelve un diccionario: column_name -> instancia column openpyxl.

    El diccionario permite trabajar usando los nombres de las columnas, en lugar de
    nombres de celdas ('A', 'B', 'C', etc...), y además asegura que el comando funcione
    aunque se cambie el orden de las columnas en el archivo .xlsx de entrada.
    """
    # Obtengo la instancia que representa la planilla
    workbook = load_workbook(filename)
    # Obtengo la pestaña activa ('active sheet')
    sheet = workbook.get_active_sheet()
    # La primera fila representa los nombres de las columnas
    column_names_cells = sheet.iter_rows(max_row=INPUT_SPREADSHEET_MAX_ROWS).next()

    columns_dict = {}
    for cell, column in izip(column_names_cells,
                             sheet.iter_cols(max_row=INPUT_SPREADSHEET_MAX_ROWS-1)):
        # Si cell.value es None quiere decir que hay que detener la iteración
        # porque no se calculó correctamente el número de columnas con datos
        if cell.value is None:
            break
        column_name = sanitize_column_name(cell.value)
        # Agrego la columna quitando la celda con el título
        columns_dict[column_name] = column[1:INPUT_SPREADSHEET_MAX_ROWS]

    validate_required_columns(columns_dict.keys())

    return columns_dict
